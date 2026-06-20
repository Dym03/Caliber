import io
import logging
import os
import tempfile
import polars as pl
from django.db import transaction
from apps.core.management.commands.init_db import parse_df, sheet_exists
from openpyxl import load_workbook
from django.db.models import Q, Subquery, OuterRef
from django.core.paginator import Paginator, PageNotAnInteger, EmptyPage
from apps.core.models import Gene, GeneVariant, PatientVariant
from apps.core.enums import ClassificationEnum
from apps.core.management.commands.init_db import (
    parse_excel_genes,
    parse_excel_dbsnp,
    parse_excel_hgvs,
)

logger = logging.getLogger(__name__)


def query_variants(
    variant: str, gene: str, dbsnp: str, page_number: int | str, per_page: int | str
) -> dict:
    gene_id = None
    if gene:
        gene_id = (
            Gene.objects.filter(symbol__iexact=gene)
            .values_list("id", flat=True)
            .first()
        )
        if not gene_id:
            return {
                "results": [],
                "meta": {"has_next": False, "total_count": 0, "total_pages": 0},
            }

    # 2. Sestavení filtrů pro klinickou DB
    patient_filters = Q()
    if gene_id:
        patient_filters &= Q(variant__genes__id=gene_id)
    if dbsnp:
        patient_filters &= Q(variant__dbsnp__iexact=dbsnp)
    if variant:
        patient_filters &= (
            Q(variant__annotations__hgvs_c__iexact=variant)
            | Q(variant__annotations__hgvs_p__iexact=variant)
            | Q(variant__variation_type__iexact=variant)
            | Q(reported_hgvs_c__iexact=variant)
        )

    patient_queryset = (
        PatientVariant.objects.select_related(
            "variant", "variant__clinvar_entry", "report", "report__patient"
        )
        .prefetch_related("variant__annotations", "variant__genes")
        .filter(patient_filters)
        .distinct()
        .order_by("-report__updated_at")
    )

    is_fallback = False
    target_queryset = patient_queryset

    # 3. Fallback na globální ClinVar katalog, pokud lokální záznamy nic nenašly
    if not patient_queryset.exists():
        is_fallback = True
        global_filters = Q()
        if gene_id:
            global_filters &= Q(genes__id=gene_id)
        if dbsnp:
            global_filters &= Q(dbsnp__iexact=dbsnp)
        if variant:
            global_filters &= (
                Q(annotations__hgvs_c__iexact=variant)
                | Q(annotations__hgvs_p__iexact=variant)
                | Q(variation_type__iexact=variant)
            )

        target_queryset = (
            GeneVariant.objects.select_related("clinvar_entry")
            .prefetch_related("annotations", "genes")
            .filter(global_filters)
            .distinct()
            .order_by("chromosome", "position")
        )

    # 4. Spuštění Paginatoru
    paginator = Paginator(target_queryset, per_page)
    try:
        page_obj = paginator.page(page_number)
    except PageNotAnInteger:
        page_obj = paginator.page(1)
    except EmptyPage:
        return {
            "results": [],
            "meta": {
                "has_next": False,
                "total_count": paginator.count,
                "total_pages": paginator.num_pages,
            },
        }

    # 5. Serializace datové struktury
    results = []
    for item in page_obj:
        var_obj = item.variant if not is_fallback else item

        transcripts = [
            f"{a.transcript_base or ''}.{a.transcript_version or ''}:{a.hgvs_c}".strip(
                ".:"
            )
            for a in var_obj.annotations.all()
        ]
        gene_symbols = [g.symbol for g in var_obj.genes.all()]

        clinvar_entry = getattr(var_obj, "clinvar_entry", None)
        clinvar_data = None
        if clinvar_entry:
            clinvar_data = {
                "id": clinvar_entry.clinvar_id,
                "score": clinvar_entry.clinvar_classification,
                "last_updated": clinvar_entry.last_updated.isoformat()
                if clinvar_entry.last_updated
                else None,
            }

        if not is_fallback:
            results.append(
                {
                    "source": "clinic",
                    "patient_id": item.report.patient.name,
                    "gene": ", ".join(gene_symbols) if gene_symbols else "Intergenic",
                    "all_genes": gene_symbols,
                    "variant": item.reported_hgvs_c
                    or (transcripts[0] if transcripts else ""),
                    "all_transcripts": transcripts,
                    "dbsnp": var_obj.dbsnp,
                    "gnomAD": var_obj.gnomAD,
                    "chromosome": var_obj.chromosome,
                    "position": var_obj.position,
                    "updated_at": item.report.updated_at.isoformat(),
                    "category": item.category or "",
                    "clinvar": clinvar_data,
                    "comment": item.comment or None,
                }
            )
        else:
            results.append(
                {
                    "source": "clinvar_catalog",
                    "patient_id": "-",
                    "gene": ", ".join(gene_symbols) if gene_symbols else "Intergenic",
                    "all_genes": gene_symbols,
                    "variant": transcripts[0] if transcripts else "Unknown Transcript",
                    "all_transcripts": transcripts,
                    "dbsnp": var_obj.dbsnp,
                    "gnomAD": var_obj.gnomAD,
                    "chromosome": var_obj.chromosome,
                    "position": var_obj.position,
                    "updated_at": None,
                    "category": "N/A",
                    "clinvar": clinvar_data,
                    "comment": "Found in reference database only.",
                }
            )

    return {
        "results": results,
        "meta": {
            "current_page": page_obj.number,
            "has_next": page_obj.has_next(),
            "has_previous": page_obj.has_previous(),
            "total_pages": paginator.num_pages,
            "total_count": paginator.count,
            "is_fallback": is_fallback,
        },
    }


def process_uploaded_variants_file(uploaded_file) -> dict:
    """
    Parses the uploaded CSV or Excel file, processes its contents, and persists the data into the database.
    Returns a dictionary containing the filename and the number of rows processed.
    """
    name = uploaded_file.name
    name_lower = name.lower()
    suffix = os.path.splitext(name_lower)[1]

    with tempfile.NamedTemporaryFile(delete=False, suffix=suffix) as tmp_file:
        for chunk in uploaded_file.chunks():
            tmp_file.write(chunk)
        tmp_path = tmp_file.name

    try:
        if suffix == ".csv":
            df = pl.read_csv(tmp_path)
        elif suffix == ".xlsx" and sheet_exists(tmp_path, "default"):
            df = pl.read_excel(tmp_path, sheet_name="default")
        else:
            try:
                df = pl.read_excel(tmp_path, sheet_name="Filtr JI")
            except Exception:
                df = pl.read_excel(tmp_path)

        with transaction.atomic():
            parse_df(df, name)

        return {"filename": name, "rows": df.height}
    finally:
        if os.path.exists(tmp_path):
            os.unlink(tmp_path)


def format_classification(clinic_val, clinvar_val) -> tuple[str, str]:
    c_str = ClassificationEnum._CLASSIFICATION_SCORE_TO_STRING.get(clinic_val)
    cv_str = ClassificationEnum._CLASSIFICATION_SCORE_TO_STRING.get(clinvar_val)

    if c_str and cv_str:
        if c_str == cv_str:
            return c_str, "Shoda (Klinika i ClinVar)"
        else:
            return (
                f"Klinika: {c_str} / ClinVar: {cv_str}",
                "Diskrepance (Neshoda hodnocení)",
            )
    elif c_str:
        return c_str, "Pouze v interní klinické DB"
    elif cv_str:
        return cv_str, "Pouze v ClinVar katalogu"
    return "—", "Nenalezeno žádné hodnocení"


def get_variant_hashes_from_db(
    unique_dbsnps: list[str], unique_variants: list[str]
) -> tuple[dict, dict]:
    """
    Queries the database for GeneVariant entries matching the provided dbSNP IDs and HGVS notations, and constructs two hash maps:
    1. dbsnp_map: Maps cleaned dbSNP IDs to their classification and ClinVar ID.
    2. hgvs_map: Maps cleaned HGVS notations (with and without gene symbol) to their classification and ClinVar ID.
    """
    dbsnp_map = {}
    hgvs_map = {}

    db_filters = Q()
    if unique_dbsnps:
        db_filters |= Q(dbsnp__in=unique_dbsnps)
    if unique_variants:
        db_filters |= Q(annotations__hgvs_c__in=unique_variants)

    if db_filters:
        latest_clinic_category = Subquery(
            PatientVariant.objects.filter(
                variant=OuterRef("pk"), category__isnull=False
            )
            .order_by("-report__updated_at")
            .values("category")[:1]
        )

        matched_variants = (
            GeneVariant.objects.select_related("clinvar_entry")
            .prefetch_related("annotations", "genes")
            .annotate(latest_clinic_cat=latest_clinic_category)
            .filter(db_filters)
            .distinct()
        )

        for gv in matched_variants:
            clinvar_score = (
                gv.clinvar_entry.clinvar_classification
                if hasattr(gv, "clinvar_entry")
                else None
            )
            clinvar_id = (
                gv.clinvar_entry.clinvar_id if hasattr(gv, "clinvar_entry") else "—"
            )
            clinic_score = gv.latest_clinic_cat

            final_score, match_status = format_classification(
                clinic_score, clinvar_score
            )

            variant_data = {
                "final_score": final_score,
                "clinvar_id": clinvar_id,
                "status": match_status,
            }

            if gv.dbsnp:
                dbsnp_map[gv.dbsnp.strip().lower()] = variant_data

            variant_genes = [g.symbol.strip().lower() for g in gv.genes.all()]

            for ann in gv.annotations.all():
                if ann.hgvs_c:
                    c_clean = ann.hgvs_c.strip()
                    # Záložní klíč bez genu pro případ prázdného variant_genes
                    hgvs_map[c_clean] = variant_data
                    for g_sym in variant_genes:
                        hgvs_map[(g_sym, c_clean)] = variant_data

    return dbsnp_map, hgvs_map


def annotate_excel_workbook(
    file_bytes: bytes, dbsnp_map: dict, hgvs_map: dict
) -> io.BytesIO:
    """
    Loads the uploaded Excel file, iterates through its rows, and annotates each variant based on dbSNP ID or HGVS notation using the provided hash maps.
    Adds three new columns for classification, ClinVar ID, and match status. Returns a Bytes
    """

    wb = load_workbook(io.BytesIO(file_bytes))
    ws = wb.active

    orig_headers = [cell.value for cell in ws[1]]
    openpyxl_gene_idx = openpyxl_variant_idx = openpyxl_dbsnp_idx = None

    for idx, h_name in enumerate(orig_headers, start=1):
        if not h_name:
            continue
        low = str(h_name).lower()
        if low in ["symbol", "gene"]:
            openpyxl_gene_idx = idx
        elif low in ["nucleotide", "hgvsc"]:
            openpyxl_variant_idx = idx
        elif low in ["vep dbsnp id", "dbsnp"]:
            openpyxl_dbsnp_idx = idx

    start_col = len(orig_headers) + 1

    ws.cell(row=1, column=start_col, value="Zhodnoceni_Patogenity")
    ws.cell(row=1, column=start_col + 1, value="ClinVar_ID")
    ws.cell(row=1, column=start_col + 2, value="Status_Kontroly")

    for row_num in range(2, ws.max_row + 1):
        gene_val = (
            ws.cell(row=row_num, column=openpyxl_gene_idx).value
            if openpyxl_gene_idx
            else None
        )
        variant_val = (
            ws.cell(row=row_num, column=openpyxl_variant_idx).value
            if openpyxl_variant_idx
            else None
        )
        dbsnp_val = (
            ws.cell(row=row_num, column=openpyxl_dbsnp_idx).value
            if openpyxl_dbsnp_idx
            else None
        )

        g_clean = parse_excel_genes(gene_val) if gene_val else ""
        d_clean = parse_excel_dbsnp(str(dbsnp_val)) if dbsnp_val else ""

        _, _, hgvs = parse_excel_hgvs(str(variant_val)) if variant_val else ""
        logger.debug(
            f"Processing row {row_num}: Gene='{gene_val}', Variant='{variant_val}', dbSNP='{dbsnp_val}' -> Cleaned Gene='{g_clean}', Cleaned Variant='{hgvs}', Cleaned dbSNP='{d_clean}'"
        )

        match = None
        if d_clean and d_clean in dbsnp_map:
            match = dbsnp_map[d_clean]
        elif g_clean and hgvs:
            for g in g_clean:
                g = g.strip().lower()
                if (g, hgvs) in hgvs_map:
                    match = hgvs_map[(g, hgvs)]
                    break

        if not match and hgvs and hgvs in hgvs_map:
            match = hgvs_map[hgvs]

        if match:
            ws.cell(row=row_num, column=start_col, value=match["final_score"])
            ws.cell(row=row_num, column=start_col + 1, value=match["clinvar_id"])
            ws.cell(row=row_num, column=start_col + 2, value=match["status"])
        else:
            ws.cell(row=row_num, column=start_col, value="Neklasifikováno")
            ws.cell(row=row_num, column=start_col + 1, value="—")
            ws.cell(row=row_num, column=start_col + 2, value="Nenalezeno v DB")

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output
