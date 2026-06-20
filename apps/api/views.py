import logging
import os
import tempfile
import io
from openpyxl import load_workbook
import polars as pl
from django.db import transaction
from django.db.models import Q, OuterRef, Subquery
from django.http import HttpResponse, JsonResponse
from django.views.decorators.http import require_POST
from django.core.paginator import Paginator, EmptyPage, PageNotAnInteger

from apps.core.enums import ClassificationEnum
from apps.core.management.commands.init_db import parse_df, parse_excel_dbsnp, parse_excel_genes, sheet_exists
from apps.core.models import Gene, GeneVariant, PatientVariant
from apps.core.management.commands.init_db import parse_excel_hgvs

logger = logging.getLogger(__name__)

def search_variants(request):
    """
    Search for variants based on query parameters.
    Query Parameters:
    - variant: The variant identifier (e.g., HGVS notation).
    - gene: The gene symbol (e.g., BRCA1).
    - dbsnp: The dbSNP identifier (e.g., rs123456).
    - page: The page number for pagination (default: 1).
    - per_page: The number of results per page (default: 50).
    Returns:
    - JSON response containing the search results and pagination metadata.

    The search first attempts to find matching variants in the internal patient records.
    If no matches are found, it falls back to searching the ClinVar database.
    """
    variant = request.GET.get("variant", "").strip()
    gene = request.GET.get("gene", "").strip()
    dbsnp = request.GET.get("dbsnp", "").strip()

    page_number = request.GET.get("page", 1)
    per_page = request.GET.get("per_page", 50)

    if not (variant or gene or dbsnp):
        return JsonResponse(
            {"results": [], "meta": {"has_next": False, "total_count": 0}}
        )

    # =========================================================================
    # STEP 1: Attempt to search internal Patient Records first
    # =========================================================================
    gene_id = None  # Simplify search by using the Gene ID directly for filtering
    if gene:
        gene_id = (
            Gene.objects.filter(symbol__iexact=gene)
            .values_list("id", flat=True)
            .first()
        )
        # Fast-fail: If they searched a gene that doesn't exist, stop immediately
        if not gene_id:
            return JsonResponse(
                {
                    "results": [],
                    "meta": {"has_next": False, "total_count": 0, "total_pages": 0},
                }
            )

    patient_filters = Q()
    if gene_id:
        patient_filters &= Q(variant__genes__id=gene_id)
    if dbsnp:
        patient_filters &= Q(variant__dbsnp__iexact=dbsnp)
    if variant:
        patient_filters &= (
            Q(variant__annotations__hgvs_c__iexact=variant)
            | Q(variant__annotations__hgvs_p__iexact=variant)
            | Q(variant__variation_type__iexact=variant)
            | Q(reported_hgvs_c__iexact=variant)
        )

    patient_queryset = (
        PatientVariant.objects.select_related(
            "variant", "variant__clinvar_entry", "report", "report__patient"
        )
        .prefetch_related("variant__annotations", "variant__genes")
        .filter(patient_filters)
        .distinct()
        .order_by("-report__updated_at")
    )

    results = []
    is_fallback = False
    target_queryset = patient_queryset

    if not patient_queryset.exists():
        is_fallback = True
        global_filters = Q()

        global_filters &= Q(genes__id=gene_id)

        if dbsnp:
            global_filters &= Q(dbsnp__iexact=dbsnp)
        if variant:
            global_filters &= (
                Q(annotations__hgvs_c__iexact=variant)
                | Q(annotations__hgvs_p__iexact=variant)
                | Q(variation_type__iexact=variant)
            )

        target_queryset = (
            GeneVariant.objects.select_related("clinvar_entry")
            .prefetch_related("annotations", "genes")
            .filter(global_filters)
            .distinct()
            .order_by(
                "chromosome", "position"
            )  # Consistent order is critical for pagination
        )

    # =========================================================================
    # STEP 2: Execute Pagination on whichever target dataset won
    # =========================================================================
    paginator = Paginator(target_queryset, per_page)

    try:
        page_obj = paginator.page(page_number)
    except PageNotAnInteger:
        page_obj = paginator.page(1)
    except EmptyPage:
        return JsonResponse(
            {"results": [], "meta": {"has_next": False, "total_count": paginator.count}}
        )

    # =========================================================================
    # STEP 3: Serialize data based on data source branch
    # =========================================================================
    for item in page_obj:
        # Resolve object based on whether we are looping over PatientVariant or GeneVariant
        var_obj = item.variant if not is_fallback else item

        transcripts = [
            f"{a.transcript_base or ''}.{a.transcript_version or ''}:{a.hgvs_c}".strip(
                ".:"
            )
            for a in var_obj.annotations.all()
        ]
        gene_symbols = [g.symbol for g in var_obj.genes.all()]

        clinvar_entry = getattr(var_obj, "clinvar_entry", None)
        clinvar_data = None
        if clinvar_entry:
            clinvar_data = {
                "id": clinvar_entry.clinvar_id,
                "score": clinvar_entry.clinvar_classification,
                "last_updated": clinvar_entry.last_updated.isoformat()
                if clinvar_entry.last_updated
                else None,
            }

        if not is_fallback:
            # Structuring patient data row
            results.append(
                {
                    "source": "clinic",
                    "patient_id": item.report.patient.name,
                    "gene": ", ".join(gene_symbols) if gene_symbols else "Intergenic",
                    "all_genes": gene_symbols,
                    "variant": item.reported_hgvs_c
                    or (transcripts[0] if transcripts else ""),
                    "all_transcripts": transcripts,
                    "dbsnp": var_obj.dbsnp,
                    "gnomAD": var_obj.gnomAD,
                    "chromosome": var_obj.chromosome,
                    "position": var_obj.position,
                    "updated_at": item.report.updated_at.isoformat(),
                    "category": item.category or "",
                    "clinvar": clinvar_data,
                    "comment": item.comment or None,
                }
            )
        else:
            # Structuring global reference catalog row
            results.append(
                {
                    "source": "clinvar_catalog",
                    "patient_id": "-",
                    "gene": ", ".join(gene_symbols) if gene_symbols else "Intergenic",
                    "all_genes": gene_symbols,
                    "variant": transcripts[0] if transcripts else "Unknown Transcript",
                    "all_transcripts": transcripts,
                    "dbsnp": var_obj.dbsnp,
                    "gnomAD": var_obj.gnomAD,
                    "chromosome": var_obj.chromosome,
                    "position": var_obj.position,
                    "updated_at": None,
                    "category": "N/A",
                    "clinvar": clinvar_data,
                    "comment": "Found in reference database only.",
                }
            )

    # Return results alongside standard pagination metadata
    return JsonResponse(
        {
            "results": results,
            "meta": {
                "current_page": page_obj.number,
                "has_next": page_obj.has_next(),
                "has_previous": page_obj.has_previous(),
                "total_pages": paginator.num_pages,
                "total_count": paginator.count,
                "is_fallback": is_fallback,
            },
        }
    )


@require_POST
def upload_variants_file(request):
    """
    Handle the upload of variant files (Excel) and process them to populate the database.
    Expects files to be uploaded under the "file" key in the request.
    Returns a JSON response indicating the success or failure of the operation.
    """
    uploaded_files = request.FILES.getlist("file")
    if not uploaded_files:
        return JsonResponse({"error": "No file provided."}, status=400)

    allowed_extensions = {".xlsx", ".xls", ".csv"}
    for uploaded in uploaded_files:
        name_lower = uploaded.name.lower()
        if not any(name_lower.endswith(ext) for ext in allowed_extensions):
            return JsonResponse(
                {"error": f"Unsupported file type: {uploaded.name}."},
                status=400,
            )

    results = []
    for uploaded in uploaded_files:
        name = uploaded.name
        name_lower = name.lower()
        suffix = os.path.splitext(name_lower)[1]
        with tempfile.NamedTemporaryFile(delete=False, suffix=suffix) as tmp_file:
            for chunk in uploaded.chunks():
                tmp_file.write(chunk)
            tmp_path = tmp_file.name

        try:
            if suffix == ".csv":
                df = pl.read_csv(tmp_path)
            elif suffix == ".xlsx" and sheet_exists(tmp_path, "default"):
                df = pl.read_excel(tmp_path, sheet_name="default")
            else:
                try:
                    df = pl.read_excel(tmp_path, sheet_name="Filtr JI")
                except Exception:
                    df = pl.read_excel(tmp_path)

            with transaction.atomic():
                parse_df(df, name)
            row_count = df.height
        finally:
            os.unlink(tmp_path)

        results.append({"filename": name, "rows": row_count})

    return JsonResponse(
        {
            "filenames": [item["filename"] for item in results],
            "rows": sum(item["rows"] for item in results),
            "files": results,
        }
    )


@require_POST
def classify_variants(request):
    """
    Handle the upload of an Excel file containing variants and classify them based on internal and ClinVar data.
    """
    if request.method != "POST":
        return JsonResponse({"error": "Requires POST method."}, status=405)

    uploaded_file = request.FILES.get("file")
    if not uploaded_file:
        return JsonResponse({"error": "No file provided."}, status=400)

    allowed_extensions = {".xlsx", ".xls", ".csv"}
    name_lower = uploaded_file.name.lower()
    if not any(name_lower.endswith(ext) for ext in allowed_extensions):
        return JsonResponse(
            {"error": f"Unsupported file type: {uploaded_file.name}."},
            status=400,
        )
    
    try:        
        file_bytes = uploaded_file.read()
        df = pl.read_excel(io.BytesIO(file_bytes))
        
        gene_col = None
        hgsv_c_col = None
        dbsnp_col = None

        for orig_col in df.columns:
            low = orig_col.lower()
            if low in ["symbol", "gene"]:
                gene_col = orig_col
            elif low in ["nucleotide"]:
                hgsv_c_col = orig_col
            elif low in ["hgvsc"]:
                hgsv_c_col = orig_col
            elif low in ["vep dbsnp id", "dbsnp"]:
                dbsnp_col = orig_col

        if not (gene_col or hgsv_c_col or dbsnp_col):
            return JsonResponse({
                "error": "Chybná struktura Excelu. Tabulka musí obsahovat alespoň jeden ze sloupců: Gen, Varianta, dbSNP."
            }, status=400)
        
        parsed_dbsnps = [
            parse_excel_dbsnp(str(x).strip())
            for x in df[dbsnp_col].drop_nulls().unique().to_list()
        ]

        unique_dbsnps = [
            db for db in parsed_dbsnps
            if db and db.strip() not in ["", "-", "—", "none", "nan"]
        ]
        
        raw_variants = [
            str(x).strip() 
            for x in df[hgsv_c_col].drop_nulls().unique().to_list() 
            if str(x).strip()
        ]
        
        unique_variants = []
        for v in raw_variants:
            _, _, hgvs = parse_excel_hgvs(v)
            if hgvs:
                unique_variants.append(hgvs)
        

        dbsnp_map = {}
        hgvs_map = {}

        db_filters = Q()
        if unique_dbsnps:
            db_filters |= Q(dbsnp__in=unique_dbsnps)
        if unique_variants:
            db_filters |= Q(annotations__hgvs_c__in=unique_variants)

        if db_filters:
            latest_clinic_category = Subquery(
                PatientVariant.objects.filter(
                    variant=OuterRef("pk"),
                    category__isnull=False
                )
                .order_by("-report__updated_at")
                .values("category")[:1]
            )

            matched_variants = (
                GeneVariant.objects.select_related("clinvar_entry")
                .prefetch_related("annotations", "genes")
                .annotate(latest_clinic_cat=latest_clinic_category)
                .filter(db_filters)
                .distinct()
            )

            def format_classification(clinic_val, clinvar_val) -> tuple[str, str]:                
                c_str = ClassificationEnum._CLASSIFICATION_SCORE_TO_STRING.get(clinic_val)
                cv_str = ClassificationEnum._CLASSIFICATION_SCORE_TO_STRING.get(clinvar_val)

                if c_str and cv_str:
                    if c_str == cv_str:
                        return c_str, "Shoda (Klinika i ClinVar)"
                    else:
                        return f"Klinika: {c_str} / ClinVar: {cv_str}", "Diskrepance (Neshoda hodnocení)"
                elif c_str:
                    return c_str, "Pouze v interní klinické DB"
                elif cv_str:
                    return cv_str, "Pouze v ClinVar katalogu"
                return "—", "Nenalezeno žádné hodnocení"
            
            for gv in matched_variants:
                clinvar_score = gv.clinvar_entry.clinvar_classification if hasattr(gv, "clinvar_entry") else None
                clinvar_id = gv.clinvar_entry.clinvar_id if hasattr(gv, "clinvar_entry") else "—"
                clinic_score = gv.latest_clinic_cat  

                final_score, match_status = format_classification(clinic_score, clinvar_score)

                variant_data = {
                    "final_score": final_score,
                    "clinvar_id": clinvar_id,
                    "status": match_status
                }

                if gv.dbsnp:
                    dbsnp_map[gv.dbsnp.strip().lower()] = variant_data

                variant_genes = [g.symbol.strip().lower() for g in gv.genes.all()]
                
                for ann in gv.annotations.all():
                    if ann.hgvs_c:
                        c_clean = ann.hgvs_c.strip()
                        # Záložní klíč bez genu pro případ prázdného variant_genes
                        hgvs_map[c_clean] = variant_data
                        for g_sym in variant_genes:
                            hgvs_map[(g_sym, c_clean)] = variant_data

        uploaded_file.seek(0)
        wb = load_workbook(io.BytesIO(file_bytes))
        ws = wb.active

        orig_headers = [cell.value for cell in ws[1]]
        openpyxl_gene_idx = openpyxl_variant_idx = openpyxl_dbsnp_idx = None

        for idx, h_name in enumerate(orig_headers, start=1):
            if not h_name:
                continue
            low = str(h_name).lower()
            if low in ["symbol", "gene"]:
                openpyxl_gene_idx = idx
            elif low in ["nucleotide", "hgvsc"]:
                openpyxl_variant_idx = idx
            elif low in ["vep dbsnp id", "dbsnp"]:
                openpyxl_dbsnp_idx = idx
        
        start_col = len(orig_headers) + 1
        
        ws.cell(row=1, column=start_col, value="Zhodnoceni_Patogenity")
        ws.cell(row=1, column=start_col + 1, value="ClinVar_ID")
        ws.cell(row=1, column=start_col + 2, value="Status_Kontroly")

        for row_num in range(2, ws.max_row + 1):
            gene_val = ws.cell(row=row_num, column=openpyxl_gene_idx).value if openpyxl_gene_idx else None
            variant_val = ws.cell(row=row_num, column=openpyxl_variant_idx).value if openpyxl_variant_idx else None
            dbsnp_val = ws.cell(row=row_num, column=openpyxl_dbsnp_idx).value if openpyxl_dbsnp_idx else None

            g_clean = parse_excel_genes(gene_val) if gene_val else ""
            d_clean = parse_excel_dbsnp(str(dbsnp_val)) if dbsnp_val else ""
            
            _, _, hgvs = parse_excel_hgvs(str(variant_val)) if variant_val else ""
            logger.info(f"Processing row {row_num}: Gene='{gene_val}', Variant='{variant_val}', dbSNP='{dbsnp_val}' -> Cleaned Gene='{g_clean}', Cleaned Variant='{hgvs}', Cleaned dbSNP='{d_clean}'")

            match = None
            if d_clean and d_clean in dbsnp_map:
                match = dbsnp_map[d_clean]
            elif g_clean and hgvs:
                for g in g_clean:
                    g = g.strip().lower()
                    if (g, hgvs) in hgvs_map:
                        match = hgvs_map[(g, hgvs)]
                        break

            if not match and hgvs and hgvs in hgvs_map:
                match = hgvs_map[hgvs]

            if match:
                ws.cell(row=row_num, column=start_col, value=match["final_score"])
                ws.cell(row=row_num, column=start_col + 1, value=match["clinvar_id"])
                ws.cell(row=row_num, column=start_col + 2, value=match["status"])
            else:
                ws.cell(row=row_num, column=start_col, value="Neklasifikováno")
                ws.cell(row=row_num, column=start_col + 1, value="—")
                ws.cell(row=row_num, column=start_col + 2, value="Nenalezeno v DB")
        
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        response = HttpResponse(
            output.getvalue(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        response["Content-Disposition"] = 'attachment; filename="analyzovaný_report.xlsx"'
        return response

    except Exception as e:
        return JsonResponse({"error": f"Chyba serveru při analýze: {str(e)}"}, status=500)