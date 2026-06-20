import logging
import re
import glob
import os
import tempfile
import polars as pl
from openpyxl import load_workbook
from django.core.management.base import BaseCommand
from django.db import transaction
from apps.core.enums import ClassificationEnum
from apps.core.models import (
    Gene,
    GeneVariant,
    GeneticReport,
    Patient,
    PatientVariant,
    TranscriptAnnotation,
)

logger = logging.getLogger(__name__)

# Dedicated in-memory caches to prevent redundant DB lookups inside the loop
patient_cache = {}
gene_cache = {}
variant_cache = {}
annotation_cache = {}
variant_gene_m2m_cache = (
    set()
)  # Tracks (variant_id, gene_id) to avoid repetitive .add() queries


def sheet_exists(path: str, sheet: str) -> bool:
    wb = load_workbook(path, read_only=True)
    return sheet in wb.sheetnames


def clean_str(value: object, null_if_empty: bool = False) -> str | None:
    if value is None or value == "-":
        return None if null_if_empty else ""
    cleaned = str(value).strip()
    return cleaned


def clean_int(value: object) -> int | None:
    if value in (None, "", "nan"):
        return None
    return int(value)


def normalize_var_type(var_type: str) -> str:
    if var_type is None:
        return ""
    var_type = var_type.lower()
    if var_type in [
        "single nucleotide variant",
        "snv",
        "snp",
        "single nucleotide polymorphism",
    ]:
        return "SNV"
    # elif var_type in ["snp", "single nucleotide polymorphism"]: // TODO Ask if this is a wanted mapping - because this kinda breaks the unique constraints for clinvar mapping
    #     return "SNP"
    elif var_type in ["deletion", "del"]:
        return "DEL"
    elif var_type in ["insertion", "ins"]:
        return "INS"
    elif var_type in ["duplication", "dup"]:
        return "DUP"
    elif var_type in ["inversion", "inv"]:
        return "INV"
    elif var_type in ["microsatellite", "repeat expansion"]:
        return "STR"
    elif var_type in ["haplotype"]:
        return "HAPLOTYPE"
    elif var_type in ["compoundheterozygous"]:
        return "COMPOUND_HET"
    elif var_type in ["indel"]:
        return "INDEL"
    else:
        logging.warning(f"Unknown variation type: {var_type}")
        return var_type.upper()


def parse_excel_hgvs(excel_string):
    match = re.match(r"(^[A-Z_]+_\d+)\.(\d+):(.*)", excel_string)
    if match:
        return match.group(1), match.group(2), match.group(3)
    return None, None, excel_string


def parse_excel_dbsnp(excel_string) -> str:
    if not excel_string:
        return ""
    # Split by commas and filter out any non-rs IDs
    dbsnp_ids = [
        id.strip() for id in excel_string.split(",") if id.strip().startswith("rs")
    ]
    if len(dbsnp_ids) > 1:
        logger.warning(
            f"Multiple dbSNP IDs found: {dbsnp_ids}. Only the first one will be used."
        )
    return dbsnp_ids[0] if dbsnp_ids else ""


def parse_excel_genes(excel_string) -> list:
    if not excel_string:
        return []
    # Split by common delimiters and filter out empty strings
    gene_symbols = [g.strip() for g in re.split(r"[,;/|]", excel_string) if g.strip()]
    return gene_symbols


def parse_row(row) -> dict:
    cleaned_data = {}
    cleaned_data["patient_id"] = clean_str(row.get("Name"))
    cleaned_data["gene_symbols"] = parse_excel_genes(
        row.get("Symbol") or row.get("Gene")
    )

    cleaned_data["variation_type"] = normalize_var_type(
        row.get("Variant_class") or row.get("Variation Type")
    )
    cleaned_data["chromosome"] = clean_str(row.get("Chr"))
    cleaned_data["position"] = clean_int(
        row.get("Coordinate") or row.get("Start Position")
    )
    cleaned_data["ref_allele"] = clean_str(row.get("Reference") or row.get("Ref"))
    cleaned_data["alt_allele"] = clean_str(row.get("Alternate") or row.get("Alt"))
    cleaned_data["dbSNP"] = parse_excel_dbsnp(
        row.get("VEP dbSNP ID", "")
        or row.get(
            "dbSNP", ""
        )  # TODO IF CMON are wanted, because right now we have some dbSNP with comma separated multiple IDs, which is not ideal for the unique constraint and the mapping to clinvar variants. We should ask if we want to split those into multiple dbSNPs or just take the first one or something else
    )

    cleaned_data["hgvs_coding"] = clean_str(row.get("HGVSc") or row.get("Transcript"))
    nucleotide = clean_str(row.get("Nucleotide", ""))
    if nucleotide:
        cleaned_data["hgvs_coding"] += f":{nucleotide}"
    cleaned_data["hgvs_p"] = clean_str(row.get("HGVSp", "") or row.get("AA Change", ""))
    cleaned_data["category"] = ClassificationEnum.from_excel_string(
        clean_str(row.get("Kategorie"))
    ).score
    cleaned_data["comment"] = clean_str(row.get("Komentář", ""))
    cleaned_data["exon"] = clean_str(row.get("Exon"))
    cleaned_data["zygosity"] = clean_str(row.get("Genotype") or row.get("Zygosity"))
    cleaned_data["gnomAD"] = clean_str(
        row.get("gnomAD AF") or row.get("gnomAD (Exome)")
    )

    (
        cleaned_data["transcript_base"],
        cleaned_data["transcript_version"],
        cleaned_data["hgvs_c"],
    ) = parse_excel_hgvs(cleaned_data["hgvs_coding"])

    return cleaned_data


def persist_row(data: dict, file_name: str):
    if data["patient_id"] not in patient_cache:
        patient, _ = Patient.objects.get_or_create(name=data["patient_id"])
        patient_cache[data["patient_id"]] = patient
    else:
        patient = patient_cache[data["patient_id"]]

    # 2. GeneVariant Lookup (Using the updated genomic coordinates unique constraint)
    variant_key = (
        data["chromosome"],
        data["position"],
        data["variation_type"],
        data["ref_allele"],
        data["alt_allele"],
    )

    if variant_key not in variant_cache:
        gene_variant, _ = GeneVariant.objects.get_or_create(
            chromosome=data["chromosome"],
            position=data["position"],
            variation_type=data["variation_type"],
            ref_allele=data["ref_allele"],
            alt_allele=data["alt_allele"],
            defaults={
                "gnomAD": data["gnomAD"],
                "dbsnp": data["dbSNP"],
            },
        )

        if not gene_variant.gnomAD and data["gnomAD"]:
            gene_variant.gnomAD = data["gnomAD"]
            gene_variant.save(update_fields=["gnomAD"])
        if not gene_variant.dbsnp and data["dbSNP"]:
            gene_variant.dbsnp = data["dbSNP"]
            gene_variant.save(update_fields=["dbsnp"])

        variant_cache[variant_key] = gene_variant
    else:
        gene_variant = variant_cache[variant_key]

    # 3. Handle ManyToMany Relationship for Genes
    # Splitting by common delimiters (like commas or slashes) in case an excel row lists overlapping genes

    for symbol in data["gene_symbols"]:
        if symbol not in gene_cache:
            gene, _ = Gene.objects.get_or_create(symbol=symbol)
            gene_cache[symbol] = gene
        else:
            gene = gene_cache[symbol]

        # Utilize the m2m cache to prevent executing repetitive SQL intermediate inserts
        m2m_key = (gene_variant.id, gene.id)
        if m2m_key not in variant_gene_m2m_cache:
            gene_variant.genes.add(gene)
            variant_gene_m2m_cache.add(m2m_key)

    # 4. TranscriptAnnotation Lookup (Moved to a dedicated annotation cache dictionary)
    annotation_key = (data["hgvs_c"], gene_variant.id)
    if annotation_key not in annotation_cache:
        annotation, _ = TranscriptAnnotation.objects.get_or_create(
            variant=gene_variant,
            hgvs_c=data["hgvs_c"],
            defaults={
                "hgvs_p": data["hgvs_p"],
                "transcript_base": data["transcript_base"],
                "transcript_version": data["transcript_version"],
                "exon": data["exon"],
            },
        )
        annotation_cache[annotation_key] = annotation

    # 5. Report & Patient Linkage
    report, _ = GeneticReport.objects.get_or_create(
        patient=patient, report_name=file_name
    )

    PatientVariant.objects.get_or_create(
        report=report,
        variant=gene_variant,
        zygosity=data["zygosity"],
        category=data["category"],
        comment=data["comment"],
        reported_hgvs_c=data["hgvs_coding"],
    )


def parse_df(df: pl.DataFrame, file_name: str):
    for row in df.iter_rows(named=True):
        cleaned_data = parse_row(row)
        persist_row(cleaned_data, file_name)


class Command(BaseCommand):
    DEFAULT_ROOT_DIR = "."

    def add_arguments(self, parser) -> None:
        parser.add_argument(
            "--root_dir",
            help="Sets the root directory for the imported xlxs files. Default is the current directory.",
            default=self.DEFAULT_ROOT_DIR,
        )

    def handle(self, *args: tuple, **options: dict) -> None:
        root_dir = options.get("root_dir") or self.DEFAULT_ROOT_DIR

        for file_name in glob.iglob(f"{root_dir}/**/*.xls*", recursive=True):
            print(f"Importing data from {file_name}...")
            df = None

            if file_name.endswith(".xlsx") and sheet_exists(file_name, "default"):
                df = pl.read_excel(file_name, sheet_name="default")
            else:
                try:
                    df = pl.read_excel(file_name, sheet_name="Filtr JI")
                except Exception as e:
                    df = pl.read_excel(file_name)

            with transaction.atomic():
                parse_df(df, file_name)
